import streamlit as st
import pandas as pd
import streamlit as st
import time   # 👈 Add this
from datetime import datetime
from zoneinfo import ZoneInfo
import io
import os
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
def indian_currency(amount):
    amount = float(amount)
    s = f"{amount:.2f}"
    integer, decimal = s.split(".")

    if len(integer) > 3:
        last3 = integer[-3:]
        rest = integer[:-3]
        parts = []
        while len(rest) > 2:
            parts.insert(0, rest[-2:])
            rest = rest[:-2]
        if rest:
            parts.insert(0, rest)
        integer = ",".join(parts) + "," + last3

    return f"₹ {integer}.{decimal}"
st.set_page_config(
    page_title="HQA E&M Open Receipt",
    layout="wide",
    initial_sidebar_state="expanded"
)
st.markdown("""
<style>
/* Hide Streamlit dataframe/table toolbar buttons (eye, download, search, fullscreen). */
div[data-testid="stElementToolbar"] {
    display: none !important;
}
</style>
""", unsafe_allow_html=True)
st.markdown("""
<style>
[data-testid="stSidebarCollapsedControl"] {
    display: hide;
}
</style>
""", unsafe_allow_html=True)
st.markdown("""
<style>
.block-container {
    max-width: 100%;
    padding-top: 1rem;
    padding-left: 1rem;
    padding-right: 1rem;
}

# -----------------------------
# Page Configuration
# -----------------------------

</style>
""", unsafe_allow_html=True)

st.markdown("""
<style>
[data-testid="stSidebar"] {
    min-width: 150px;
    max-width: 150px;
}
</style>
""", unsafe_allow_html=True)


# Logo
st.image("tasl_logo.png", width=300)

# -----------------------------
# Title
# -----------------------------

st.title("📊 HQA E&M Open Recepite Dashboard")

# -----------------------------
# Upload Excel
# -----------------------------
uploaded_file = st.file_uploader(
    "📂 Upload MB52 Excel File",
    type=["xlsx", "xls"]
)


if uploaded_file is None:
    st.info("Please upload an MB52 Excel file.")
    st.stop()


# -----------------------------
# Read Excel
# -----------------------------
@st.cache_data
def load_excel(uploaded_file):
    return pd.read_excel(uploaded_file)

start = time.perf_counter()

try:
    df = load_excel(uploaded_file)
    st.write("Read Excel:", round(time.perf_counter() - start, 2), "sec")
except Exception as e:
    st.error(f"Unable to read Excel file.\n\n{e}")
    st.stop()

if df.empty:
    st.error("Uploaded file is empty.")
    st.stop()

# =====================================================
# SAP MB52 - READ COLUMNS BY HEADER NAME
# =====================================================

def clean_header(name):
    return (
        str(name)
        .replace("\n", " ")
        .replace("\r", " ")
        .strip()
        .lower()
    )


# Clean Excel headers
df.columns = [clean_header(col) for col in df.columns]


def find_column(possible_names):
    for col in df.columns:
        if col in [clean_header(x) for x in possible_names]:
            return col

    return None


# -----------------------------
# Find SAP columns
# -----------------------------

material_col = find_column([
    "Material"
])

description_col = find_column([
    "Material Description"
])

plant_col = find_column([
    "Plant"
])

grn_col = find_column([
    "GRN",
    "GRN No",
    "GRN Number"
])

grn_date_col = find_column([
    "GRN Date",
    "GRN Posting Date",
    "Posting Date"
])

qty_col = find_column([
    "Quality inspection",
    "Quality Inspection",
    "Quality inspection Qty",
    "Quality Inspection Qty"
])

value_col = find_column([
    "Value in QualInsp.",
    "Value in QualInsp",
    "Value in Quality Inspection"
])


# -----------------------------
# Check required columns
# -----------------------------

required = {
    "Material": material_col,
    "Material Description": description_col,
    "Plant": plant_col,
    "GRN": grn_col,
    "GRN Date": grn_date_col,
    "Quality inspection": qty_col,
    "Value": value_col
}

missing = [
    name
    for name, column in required.items()
    if column is None
]

if missing:
    st.error(
        "SAP Excel format is incorrect.\n\n"
        "Missing column(s):\n"
        + "\n".join("• " + x for x in missing)
    )

    st.write("Actual SAP headers found:")
    st.write(df.columns.tolist())

    st.stop()


# -----------------------------
# Create standard app columns
# -----------------------------

df["Material"] = (
    df[material_col]
    .fillna("")
    .astype(str)
    .str.replace(".0", "", regex=False)
    .str.strip()
)

df["Material Description"] = (
    df[description_col]
    .fillna("")
    .astype(str)
    .str.strip()
)

df["Plant"] = (
    df[plant_col]
    .fillna("")
    .astype(str)
    .str.replace(".0", "", regex=False)
    .str.strip()
)

df["GRN"] = (
    df[grn_col]
    .fillna("")
    .astype(str)
    .str.replace(".0", "", regex=False)
    .str.strip()
)

df["GRN DATE"] = pd.to_datetime(
    df[grn_date_col],
    errors="coerce"
)

df["Qty"] = pd.to_numeric(
    df[qty_col],
    errors="coerce"
).fillna(0)

df["Qty"] = df["Qty"].map(
    lambda x: f"{x:g}"
)

df["Value"] = pd.to_numeric(
    df[value_col],
    errors="coerce"
).fillna(0)

# -----------------------------
# Prepare Data
# -----------------------------
df["Material"] = (
  df["Material"]
    .astype(str)
    .str.replace(".0", "", regex=False)
    .str.strip()
)

df["Plant"] = (
    df["Plant"]
    .fillna("")
    .apply(lambda x: str(x).split(".")[0])
    .str.strip()
)

df["GRN"] = (
    df["GRN"]
    .fillna("")
    .apply(lambda x: str(x).split(".")[0])
    .str.strip()
)

df["GRN DATE"] = pd.to_datetime(
    df["GRN DATE"],
    errors="coerce"
)

df["Value"] = pd.to_numeric(
    df["Value"],
    errors="coerce"
).fillna(0)

df["Material Description"] = (
    df["Material Description"]
    .fillna("")
    .astype(str)
    .str.strip()
)

df["Qty"] = pd.to_numeric(
    df["Qty"],
    errors="coerce"
).fillna(0)

df["Qty"] = df["Qty"].map(
    lambda x: f"{x:g}"
)

# ============================================================
# DEPARTMENT CLASSIFICATION
# ============================================================

def load_master():
    return pd.read_excel(
        "material_master.xlsx",
        header=2
    )

master = load_master()

def normalize_material(value):
    if pd.isna(value):
        return ""

    value = str(value).strip()

    if value.endswith(".0"):
        value = value[:-2]

    return value

# Column A = Material
electrical_materials = set(
    master.iloc[:, 0].apply(normalize_material)
)

electrical_materials.discard("")

# -----------------------------------------------------
# Mechanical item finding
# -----------------------------------------------------
def get_department(material):
    material = str(material).strip()

    # 3852 series is Electrical
    if material.startswith("3852"):
        return "Electrical"

    # Mechanical starting series
    mechanical_prefixes = (
        "1091",
        "1092",
        "1093",
        "1094",
        "1100",
        "2",
        "3",
        "5"
    )

    if material.startswith(mechanical_prefixes):
        return "Mechanical"

# -----------------------------------------------------
# Mechanical descriptions for Series 4
# -----------------------------------------------------

series4_mechanical_keywords = [
    "HEAT SHRINKABLE TUBE",
    "HEAT SHRINK",
    "BOOT",
    "SLEEVE"
]


# -----------------------------------------------------
# Department Logic
# -----------------------------------------------------

def get_department(row):

    material = normalize_material(row["Material"])

    description = str(
        row["Material Description"]
    ).upper().strip()

    # 1. Electrical Material Master
    if material in electrical_materials:
        return "Electrical"

    # 2. Mechanical starting series
    mechanical_prefixes = (
        "1091",
        "1092",
        "1093",
        "1094",
        "1100",
        "2",
        "3",
        "5"
    )

    if material.startswith(mechanical_prefixes):
        return "Mechanical"

    # 3. Series 4 special Mechanical items
    series4_mechanical_keywords = [
        "HEAT SHRINKABLE TUBE",
        "HEAT SHRINKABLE BOOT",
        "HEAT SHRINKABLE SLEEVE"
    ]

    if material.startswith("4"):

        if any(
            keyword in description
            for keyword in series4_mechanical_keywords
        ):
            return "Mechanical"

        # Other 4-series items are Electrical
        return "Electrical"

    # 4. Not identified
    return "New Item"


# Apply Department
df["Department"] = df.apply(
    get_department,
    axis=1
)


# -----------------------------------------------------
# Apply
# -----------------------------------------------------

df["Department"] = df.apply(
    get_department,
    axis=1
)

# -----------------------------
# Department Logic
# -----------------------------

# Convert Plant codes to names
plant_map = {
    "1201": "Ecity",
    "1202": "Vemgal"
}

df["Plant"] = (
    df["Plant"]
      .astype(str)
      .str.strip()
      .map(plant_map)
      .fillna(df["Plant"].astype(str))
)

# -----------------------------
# Sidebar Filters
# -----------------------------
st.sidebar.markdown(
    f"""
    <p style="font-size:14px; margin-bottom:5px;">
    📅 {datetime.now(ZoneInfo("Asia/Kolkata")).strftime('%d-%m-%Y')}<br>
    """,
    unsafe_allow_html=True
)

st.sidebar.header("Filters")

plant_list = ["All Plants"] + sorted(
    df["Plant"].dropna().unique().tolist()
)

dept_list = [
    "All",
    "Electrical",
    "Mechanical"
]

selected_plant = st.sidebar.selectbox(
    "🏭 Plant",
    plant_list
)

selected_department = st.sidebar.selectbox(
    "⚙️ Department",
    dept_list
)
#Tasl Logo
st.sidebar.markdown("---")

st.sidebar.image("./tasl_logo.png", width=300)

# -----------------------------
# Apply Filters
# -----------------------------
filtered = df.copy()

if selected_plant != "All Plants":
    filtered = filtered[
        filtered["Plant"] == selected_plant
    ]

if selected_department != "All":
    filtered = filtered[
        filtered["Department"] == selected_department
    ]

# =====================================================
# KPI CARDS
# =====================================================
st.markdown("---")
st.subheader("📌 Dashboard")

kpi1, kpi2, kpi3 = st.columns(3)

total_value = filtered["Value"].sum()
grn_count = filtered["GRN"].nunique()
lot_count = len(filtered)

kpi1.metric(
    "💰 Pending Value",
    indian_currency(total_value)
)

kpi2.metric(
    "📄 GRN Count",
    grn_count
)

kpi3.metric(
    "📦 Lot Count",
    lot_count
)

# =====================================================
# SUMMARY TABLES
# =====================================================

# Department Pending Value
dept_summary = (
    filtered
    .groupby("Department", as_index=False)
    .agg(Pending_Value=("Value", "sum"))
)

# Department GRN & Lot Count
dept_count = (
    filtered
    .groupby("Department", as_index=False)
    .agg(
        GRN_Count=("GRN", "nunique"),
        Lot_Count=("GRN", "count")
    )
)

# Plant Summary
summary = (
    filtered[
        filtered["Plant"].notna() &
        (filtered["Plant"].astype(str).str.strip() != "")
    ]
    .groupby(["Plant", "Department"], as_index=False)
    .agg(
        GRN_Count=("GRN", "nunique"),
        Lot_Count=("GRN", "count"),
        Pending_Value=("Value", "sum")
    )
)

st.markdown("---")

col1, col2 = st.columns(2)

with col1:
    st.subheader("⚙️ Department Pending Value")
    dept_summary_display = dept_summary.copy()
    dept_summary_display["Pending_Value"] = dept_summary_display["Pending_Value"].apply(indian_currency)
    dept_summary_display.rename(columns={"Pending_Value": "Pending Value"}, inplace=True)
    st.dataframe(
        dept_summary_display,
        hide_index=True,
        use_container_width=True
    )

with col2:
    st.subheader("📄 Department GRN & Lot Count")
    st.dataframe(
        dept_count,
        hide_index=True,
        use_container_width=True
    )

st.subheader("📊 Plant Summary")

if selected_plant == "All Plants":
    plant_summary = summary
else:
    plant_summary = summary[summary["Plant"] == selected_plant]

table_height = 35 * (len(plant_summary) + 1) + 5

plant_summary_display = plant_summary.copy()
plant_summary_display["Pending_Value"] = plant_summary_display["Pending_Value"].apply(indian_currency)
plant_summary_display.rename(columns={"Pending_Value": "Pending Value"}, inplace=True)

st.dataframe(
    plant_summary_display,
    hide_index=True,
    width=430,
    height=table_height
)

display_df = filtered.copy()
display_df = filtered.copy()

# =====================================================
# DETAILED DATA
# =====================================================

st.markdown("---")
st.subheader("📋 Detailed Pending Data")

# Create detail_df FIRST
detail_df = display_df.copy()

# Filters
col1, col2, col3, col4= st.columns(4)

with col1:
    detail_plant = st.selectbox(
        "🏭 Plant",
        ["All"] + sorted(detail_df["Plant"].unique().tolist())
    )

with col2:
    detail_department = st.selectbox(
        "⚙️ Department",
        ["All"] + sorted(detail_df["Department"].unique().tolist())
    )

with col3:
    detail_grn = st.multiselect(
        "📄 Search GRN",
        options=sorted(detail_df["GRN"].astype(str).unique().tolist()),
        placeholder="Type or select one or more GRNs..."
    )
with col4:
    detail_grn_date = st.selectbox(
        "📅 GRN Date",
        ["All"] + sorted(
            detail_df["GRN DATE"]
            .dropna()
            .dt.strftime("%d-%m-%Y")
            .unique()
            .tolist()
        )
    )

filtered_detail = detail_df.copy()


if detail_plant != "All":
    filtered_detail = filtered_detail[
        filtered_detail["Plant"] == detail_plant
    ]

if detail_department != "All":
    filtered_detail = filtered_detail[
        filtered_detail["Department"] == detail_department
    ]

if detail_grn:
    filtered_detail = filtered_detail[
        filtered_detail["GRN"].astype(str).isin(detail_grn)
    ]
if detail_grn_date != "All":
    filtered_detail = filtered_detail[
        filtered_detail["GRN DATE"].dt.strftime("%d-%m-%Y") == detail_grn_date
    ]
filtered_detail = filtered_detail[
    filtered_detail["Plant"].fillna("").astype(str).str.strip() != ""
]
filtered_detail["Lot Pending"] = (
    filtered_detail.groupby("GRN")["GRN"].transform("size")
)

detail = (
    filtered_detail.groupby(
        [
            "Plant",
            "Department",
            "GRN",
            "GRN DATE"
        ],
        as_index=False
    ).agg(
        Lot_Pending=("GRN", "size"),
        Value=("Value", "sum")
    )
)

from pandas.tseries.offsets import BDay
import pandas as pd

# Convert GRN DATE to datetime
detail["GRN DATE"] = pd.to_datetime(
    detail["GRN DATE"],
    errors="coerce"
)
# Today's date
today = pd.Timestamp(datetime.now(ZoneInfo("Asia/Kolkata")).date())

# Calculate 5 working day due date (Saturday & Sunday excluded)
detail["Due Date"] = detail["GRN DATE"] + BDay(4)

# Calculate 5 working day due date (Saturday & Sunday excluded)
detail["Due Date"] = detail["GRN DATE"] + BDay(4)

# Quarter-end override
quarter_end = detail["GRN DATE"] + pd.offsets.QuarterEnd(0)

mask = (
    detail["GRN DATE"].dt.month.isin([3, 6, 9, 12]) &
    (detail["Due Date"] > quarter_end)
)

detail.loc[mask, "Due Date"] = quarter_end[mask]

# Calculate Ageing
today = pd.Timestamp.today().normalize()

detail["Closing 4 Days"] = (
    detail["Due Date"] - today
).dt.days

detail["GRN DATE"] = detail["GRN DATE"].dt.strftime("%d-%m-%Y")

# Show Due Date without time
detail["Due day"] = detail["Due Date"].dt.strftime("%d-%m-%Y")

# Remove .000000 from GRN

detail["GRN"] = (

    pd.to_numeric(detail["GRN"], errors="coerce")

    .astype("Int64")

    .astype(str)

)

# =====================================================
# Remove  .00000 from value
# =====================================================

detail["Value"] = detail["Value"].apply(indian_currency)

detail.rename(columns={
    "Plant": "Plant",
    "Department": "Department",
    "GRN": "GRN",
    "GRN DATE": "GRN Date",
    "Due Date": "Closing Date",
    "Closing 4 Days": "Days Left",
    "Value": "Pending Value (₹)",
    "Lot_Pending": "Lot Pending"
}, inplace=True)

# Highlight overdue rows
def highlight_overdue(row):
    if row["Days Left"] < 0:
        return ["background-color: #ffcccc"] * len(row)
    return [""] * len(row)
    
# Rename column
detail.rename(columns={"value": "Value"}, inplace=True)

# Remove Closing Date column
detail_display = detail.drop(columns=["Closing Date"], errors="ignore")

# Keep the detailed table in the requested order.
requested_columns = [
    "Plant", "Department", "GRN", "GRN Date",
    "Lot Pending", "Pending Value (₹)", "Days Left", "Due day"
]
detail_display = detail_display[[c for c in requested_columns if c in detail_display.columns]]

# =====================================================
# GRN Details
# =====================================================

st.markdown("---")
st.subheader("🔍 GRN Details")

# Select a GRN from the Detailed Pending Data table above.
# If row selection is supported, clicking a row opens its GRN details.
selected_grn = None
try:
    event = st.dataframe(
        detail_display.style.apply(highlight_overdue, axis=1),
        use_container_width=True,
        hide_index=True,
        on_select="rerun",
        selection_mode="single-row",
        key="detailed_pending_table",
    )
    selected_rows = event.selection.rows if event and hasattr(event, "selection") else []
    if selected_rows:
        selected_grn = str(detail_display.iloc[selected_rows[0]]["GRN"]).strip()
except TypeError:
    # Compatibility with older Streamlit versions.
    st.dataframe(
        detail_display.style.apply(highlight_overdue, axis=1),
        use_container_width=True,
        hide_index=True
    )

if selected_grn is None:
    st.info("👆 Select a GRN row above to view all materials in that GRN.")
else:
    # All source rows belonging to the selected GRN.
    grn_source = filtered[
        filtered["GRN"].astype(str).str.strip() == selected_grn
    ].copy()

    # Count source rows for each material within this GRN.
    material_count = (
        grn_source.groupby("Material", dropna=False)
        .size()
        .rename("Lot")
        .reset_index()
    )

    grn_details = grn_source[[
        "GRN",
        "Material",
        "Material Description"
    ]].copy()

    # One row per GRN + Material + Description, with the number of
    # source rows for that material shown in the Lot column.
    grn_details = grn_details.drop_duplicates(
        subset=["GRN", "Material", "Material Description"]
    )

    grn_details = grn_details.merge(
        material_count,
        on="Material",
        how="left"
    )

    grn_details["Lot"] = grn_details["Lot"].fillna(0).astype(int)

    grn_details = grn_details[[
        "GRN", "Material", "Material Description", "Lot"
    ]]

    st.dataframe(
        grn_details,
        use_container_width=True,
        hide_index=True
    )


#Excel Download
output = io.BytesIO()

with pd.ExcelWriter(output, engine="openpyxl") as writer:

    electrical = detail[detail["Department"] == "Electrical"]
    mechanical = detail[detail["Department"] == "Mechanical"]

    electrical.to_excel(
        writer,
        sheet_name="Electrical",
        index=False
    )

    mechanical.to_excel(
        writer,
        sheet_name="Mechanical",
        index=False
    )

    workbook = writer.book

    header_fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for ws in workbook.worksheets:

        last_row = ws.max_row + 1

        ws.cell(last_row, 1).value = "TOTAL"

        value_col = ws.max_column

        ws.cell(
            last_row,
            value_col
        ).value = f"=SUM({get_column_letter(value_col)}2:{get_column_letter(value_col)}{last_row-1})"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        ws.auto_filter.ref = ws.dimensions

        # Indian number format
for row in ws.iter_rows(min_row=2):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            if float(cell.value).is_integer():
                cell.value = int(cell.value)
            cell.number_format = '#,##,##0.00'

        for col in ws.columns:
            length = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = length + 3

output.seek(0)

st.download_button(
    "📥 Download Excel",
    output,
    file_name="Pending_Report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
# =====================================================
# DOWNLOAD SUMMARY
# =====================================================

st.markdown("---")

download_df = summary.copy()

excel_data = download_df.to_csv(index=False).encode("utf-8")

st.download_button(
    label="📥 Download Summary",
    data=excel_data,
    file_name=f"MB52_Pending_{datetime.now().strftime('%d%m%Y')}.csv",
    mime="text/csv"
)

# =====================================================
# DOWNLOAD DETAIL (EXCEL)
# =====================================================

import io
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# -----------------------------------------------------
# Build the requested detailed download report
# -----------------------------------------------------

def source_col(aliases):
    """Find a source Excel column using the same cleaned-header logic."""
    for alias in aliases:
        cleaned = clean_header(alias)
        for col in df.columns:
            if clean_header(col) == cleaned:
                return col
    return None


def source_series(frame, aliases, default=""):
    col = source_col(aliases)
    if col is None:
        return pd.Series([default] * len(frame), index=frame.index)
    return frame[col]

# Use the currently filtered data so the download matches the dashboard filters.
report = filtered.copy()

# Calculate Due Day for every source row using the same dashboard rule.
report_grn_date = pd.to_datetime(report["GRN DATE"], errors="coerce")
report_due_date = report_grn_date + BDay(4)
report_quarter_end = report_grn_date + pd.offsets.QuarterEnd(0)
report_mask = (
    report_grn_date.dt.month.isin([3, 6, 9, 12]) &
    (report_due_date > report_quarter_end)
)
report_due_date.loc[report_mask] = report_quarter_end.loc[report_mask]

report_days_left = (
    report_due_date - pd.Timestamp.now().normalize()
).dt.days

# Requested column order.
report_columns = pd.DataFrame(index=report.index)
report_columns["Plant"] = report["Plant"]
report_columns["Department"] = report["Department"]
report_columns["GRN Date"] = report_grn_date
report_columns["Due Day"] = report_due_date
report_columns["Days Left"] = report_days_left
report_columns["GRN"] = report["GRN"]
report_columns["Material"] = report["Material"]
report_columns["Material Description"] = report["Material Description"]
report_columns["Qty"] = pd.to_numeric(report["Qty"], errors="coerce").fillna(0)
report_columns["Base unit of measure"] = source_series(
    report,
    ["Base Unit of Measure", "Base Unit", "Base UoM", "Base unit of measure"]
)
report_columns["Value in Qualinsp."] = pd.to_numeric(
    report["Value"], errors="coerce"
).fillna(0)
report_columns["Currency"] = source_series(
    report, ["Currency", "Currency Key", "Curr."]
)
report_columns["Storage location"] = source_series(
    report, ["Storage Location", "Storage Loc.", "SLoc", "SLoc."]
)
report_columns["Descr. of storage loc."] = source_series(
    report,
    [
        "Descr. of Storage Loc.",
        "Description of Storage Location",
        "Storage Location Description",
        "Descr of storage loc."
    ]
)
report_columns["Wbs element"] = source_series(
    report, ["WBS Element", "WBS", "Wbs element"]
)
report_columns["Material group"] = source_series(
    report, ["Material Group", "Material Grp", "Mat. Group"]
)
report_columns["Material type"] = source_series(
    report, ["Material Type", "Material Typ.", "Mat. Type"]
)
report_columns["Po no"] = source_series(
    report, ["PO No", "PO No.", "PO Number", "Purchase Order", "Purchasing Document"]
)

# Keep dates as real Excel dates; Excel formatting below removes 00:00:00.
report_columns["GRN Date"] = pd.to_datetime(report_columns["GRN Date"], errors="coerce")
report_columns["Due Day"] = pd.to_datetime(report_columns["Due Day"], errors="coerce")

# Clean text fields.
for col in report_columns.columns:
    if col not in ["GRN Date", "Due Day", "Qty", "Value in Qualinsp."]:
        report_columns[col] = report_columns[col].fillna("").astype(str).str.strip()

# GRN and Material should not contain Excel-style .0.
for col in ["GRN", "Material"]:
    report_columns[col] = (
        report_columns[col]
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
    )

# Excel report.
output = io.BytesIO()

with pd.ExcelWriter(output, engine="openpyxl") as writer:
    report_columns.to_excel(
        writer,
        sheet_name="Detailed Data",
        index=False
    )

    ws = writer.sheets["Detailed Data"]

    # Header: blue fill + bold white font.
    header_fill = PatternFill(
        start_color="4472C4",
        end_color="4472C4",
        fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Filter on the complete data range.
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Date format: no 00:00:00.
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=3).number_format = "dd-mm-yyyy"  # GRN Date
        ws.cell(row=row, column=4).number_format = "dd-mm-yyyy"  # Due Day
        ws.cell(row=row, column=8).number_format = "0.##"       # Qty
        ws.cell(row=row, column=10).number_format = "[$-en-IN]#,##,##0.00" # Value

    # Highlight rows that are 5 or more days overdue in red.
    # Days Left is calculated from Due Day and today's date.
    red_fill = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )
    red_font = Font(color="9C0006")

    for excel_row, days_left in enumerate(report_days_left.tolist(), start=2):
        if pd.notna(days_left) and float(days_left) <= -5:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=excel_row, column=col).fill = red_fill
                ws.cell(row=excel_row, column=col).font = red_font

    # Indian number formatting for Value in QualInsp.
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=10).number_format = "[$-en-IN]#,##,##0.00"

    # Readable column widths.
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        ws.column_dimensions[letter].width = min(max(max_length + 3, 12), 35)

output.seek(0)

st.download_button(
    label="📥 Download Detailed Data",
    data=output,
    file_name=f"HQA_EM_Detailed_Data_{datetime.now().strftime('%d%m%Y')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
# =====================================================
# Pending Status Mail 
# =====================================================

# Calculate totals from Department Summary
mech_total = dept_summary.loc[
    dept_summary["Department"] == "Mechanical",
    "Pending_Value"
].sum()

elec_total = dept_summary.loc[
    dept_summary["Department"] == "Electrical",
    "Pending_Value"
].sum()

total_value = dept_summary["Pending_Value"].sum()

today = pd.Timestamp.today()

mail_text = f"""
Dear Sir,

Please find below the HQA Open Receipt pending value as of {today.strftime('%d-%b-%Y')}.

Mechanical Department: {indian_currency(mech_total)}
Electrical Department: {indian_currency(elec_total)}

Total Pending Value: {indian_currency(total_value)}

Regards,
HQA Team.
"""

st.subheader("📧 Mail Content")

st.markdown("""
<style>
textarea {
    font-size:17px !important;
    font-family:Calibri, Arial, sans-serif !important;
}
</style>
""", unsafe_allow_html=True)

st.text_area(
    "📧 Copy and paste into Outlook",
    value=mail_text,
    height=350,
)
# =====================================================
# FOOTER
# =====================================================

st.markdown("---")

st.caption(
    f"""
HQA E&M Open Receipt Pending Dashboard

Records : {len(display_df):,}

Generated : {datetime.now(ZoneInfo("Asia/Kolkata")).strftime('%d-%m-%Y %H:%M:%S')}
"""
)
